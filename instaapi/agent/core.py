"""
InstaAgent — Core AI Agent
===========================
Main agent class that connects LLM + InstaAPI + Executor + Permissions.

Usage:
    from instaapi import Instagram
    from instaapi.agent import InstaAgent, Permission

    ig = Instagram.from_env(".env")

    # With OpenAI
    agent = InstaAgent(ig, provider="openai", api_key="sk-...")

    # With Gemini
    agent = InstaAgent(ig, provider="gemini", api_key="...")

    # Ask a question
    result = agent.ask("How many followers does Cristiano have?")
    print(result)

    # Interactive chat
    agent.chat()

    # Full access (no permission prompts)
    agent = InstaAgent(ig, provider="gemini", api_key="...",
                       permission=Permission.FULL_ACCESS)
"""

import json
import logging
import os
import time
from dataclasses import dataclass, field
from typing import Any, Callable, Dict, List, Optional

from .knowledge import SYSTEM_PROMPT
from .executor import SafeExecutor, ExecutionResult
from .permissions import Permission, PermissionManager, ActionType
from .providers import get_provider, resolve_api_key
from .providers.base import BaseProvider, ProviderResponse, ToolCall
from .tools import TOOL_HANDLERS
from .memory import AgentMemory
from .plugins import PluginManager
from .cost_tracker import CostTracker
from .retry import RetryPolicy
from .streaming import StreamHandler
from .compat import (
    safe_print, emoji, setup_console_encoding,
    get_default_memory_dir, get_default_cost_path,
    get_platform_info,
)

logger = logging.getLogger("instaapi.agent")


@dataclass
class AgentResult:
    """Result of an agent operation."""
    answer: str = ""
    code_executed: str = ""
    execution_result: Optional[ExecutionResult] = None
    files_created: List[str] = field(default_factory=list)
    steps: int = 0
    tokens_used: int = 0
    duration: float = 0.0
    error: str = ""

    @property
    def success(self) -> bool:
        return not self.error

    def __str__(self) -> str:
        if self.error:
            return f"❌ {self.error}"
        return self.answer or "Completed"


class InstaAgent:
    """
    AI Agent for InstaAPI.

    Connects any LLM with InstaAPI library.
    User speaks naturally, agent writes and executes code.

    Supported modes:
        - Login: Instagram.from_env(".env") — full API access
        - Anonymous: Instagram() or ig=None — only ig.public.*
        - Async: AsyncInstagram — async wrapper

    Supported providers:
        openai, gemini, claude, deepseek, qwen, groq, together,
        mistral, ollama, openrouter, fireworks, perplexity, xai, custom

    Args:
        ig: Instagram instance (None = anonymous mode)
        provider: AI provider name (default: 'gemini')
        api_key: AI provider API key (or auto-detect from env)
        model: Optional model override
        permission: Permission level (ASK_EVERY, ASK_ONCE, FULL_ACCESS)
        permission_callback: Custom permission prompt function
        max_steps: Maximum agent loop iterations (default 15)
        timeout: Code execution timeout in seconds (default 30)
        verbose: Print step-by-step progress
    """

    def __init__(
        self,
        ig=None,
        provider: str = "gemini",
        api_key: Optional[str] = None,
        model: Optional[str] = None,
        permission: Permission = Permission.ASK_EVERY,
        permission_callback: Optional[Callable] = None,
        max_steps: int = 15,
        timeout: int = 30,
        verbose: bool = True,
        memory: bool = False,
        memory_dir: Optional[str] = None,
        cost_tracking: bool = True,
        retry_count: int = 2,
        streaming: bool = False,
    ):
        # Resolve API key
        api_key = api_key or resolve_api_key(provider)
        if not api_key and provider.lower() != "ollama":
            raise ValueError(
                f"API key not found for '{provider}'!\n"
                f"  1. Pass api_key='...' parameter\n"
                f"  2. Add the key to .env file\n"
                f"  Supported: openai, gemini, claude, deepseek, "
                f"qwen, groq, together, mistral, ollama, openrouter, xai"
            )

        self._ig = ig
        self._provider: BaseProvider = get_provider(provider, api_key, model)
        self._model_name = model or self._provider.model_name if hasattr(self._provider, 'model_name') else 'auto'
        self._executor = SafeExecutor(ig_instance=ig, timeout=timeout)
        self._permissions = PermissionManager(
            level=permission,
            prompt_callback=permission_callback,
        )
        self._max_steps = max_steps
        self._verbose = verbose
        self._history: List[Dict[str, Any]] = []
        self._files_created: List[str] = []

        # User data cache — persists across agent calls
        self._user_cache: Dict[str, Any] = {}

        # Advanced features
        self._memory = AgentMemory(memory_dir=memory_dir) if memory else None
        self._plugins = PluginManager()
        self._cost_tracker = CostTracker() if cost_tracking else None
        self._retry = RetryPolicy(max_retries=retry_count)
        self._stream = StreamHandler(mode="cli") if streaming else None
        self._session_id = AgentMemory.generate_session_id() if memory else ""

        # Detect mode — check if actually logged in
        self._is_logged_in = False
        ig_class = type(ig).__name__ if ig else "None"
        if ig is None:
            self._mode = "anonymous"
        elif "Async" in ig_class:
            self._mode = "async"
            self._is_logged_in = True
        else:
            self._mode = "sync"
            # Check if ig actually has sessions (cookies)
            try:
                if hasattr(ig, '_session_mgr') and ig._session_mgr.session_count > 0:
                    self._is_logged_in = True
                elif hasattr(ig, '_client') and ig._client is not None:
                    self._is_logged_in = True
                else:
                    self._is_logged_in = False
            except Exception:
                self._is_logged_in = False

        # Inject cache and login status into executor
        self._executor._user_cache = self._user_cache
        self._executor._is_logged_in = self._is_logged_in

        # Build system prompt with detailed mode info
        self._is_local_provider = provider.lower() in ("ollama", "lmstudio", "local", "lm-studio", "lm_studio", "llmstudio")

        if self._is_local_provider:
            # Compact prompt for local models (limited context window)
            system_content = self._build_compact_prompt()
        else:
            mode_info = self._build_mode_info()
            # Dynamic API reference — auto-discover available modules/methods
            api_ref = self._build_api_reference()
            system_content = SYSTEM_PROMPT + mode_info + api_ref

        # Initialize with system prompt
        self._history.append({
            "role": "system",
            "content": system_content,
        })

        mode_emoji = {"sync": "🔑", "anonymous": "👤", "async": "⚡"}
        logger.info(
            f"Agent ready | Provider: {self._provider.provider_name} | "
            f"Mode: {self._mode} {mode_emoji.get(self._mode, '')} | "
            f"Permission: {permission.value}"
        )

        if self._verbose:
            setup_console_encoding()
            mode_text = {"sync": "Login", "anonymous": "Anonymous", "async": "Async"}
            safe_print(
                f"{emoji('🤖', '[BOT]')} Agent: {self._provider.provider_name} | "
                f"Mode: {mode_text.get(self._mode, self._mode)}"
            )

    # ═══════════════════════════════════════════════════════════
    # MODE DETECTION & SMART CONTEXT
    # ═══════════════════════════════════════════════════════════

    def _build_compact_prompt(self) -> str:
        """Build a compact system prompt for local models with limited context."""
        mode = "anonymous" if not self._is_logged_in else "logged_in"
        parts = [
            "You are InstaAPI Agent — an Instagram analytics assistant.",
            "You write Python code using the `ig` object to interact with Instagram.",
            "",
            "RULES:",
            "1. Write executable Python code in ```python blocks",
            "2. Use try/except for all API calls",
            "3. Use f-strings for formatted output",
            "4. Cache results in `_cache` dict",
            "5. NEVER use nested quotes in f-strings",
            "",
        ]

        if mode == "anonymous":
            parts.extend([
                "MODE: ANONYMOUS — only ig.public.* methods work",
                "Main methods:",
                "- ig.public.get_profile(username) → dict",
                "- ig.public.get_posts(username, max_count=12) → list",
                "- ig.public.get_post_by_url(url) → dict",
                "- ig.public.get_comments(shortcode) → list",
                "- ig.public.get_hashtag_posts(tag) → list",
                "",
                "Profile fields: username, biography, is_verified, is_private,",
                "  edge_followed_by.count (followers), edge_follow.count (following),",
                "  edge_owner_to_timeline_media.count (posts), profile_pic_url_hd",
                "",
                "Example:",
                "```python",
                "profile = ig.public.get_profile('username')",
                "if profile:",
                "    followers = profile.get('edge_followed_by', {}).get('count', 0)",
                "    print(f'Followers: {followers:,}')",
                "```",
            ])
        else:
            parts.extend([
                "MODE: LOGGED IN — all API methods available",
                "Main methods:",
                "- ig.users.get_by_username(username) → User model",
                "- ig.public.get_profile(username) → dict (fallback)",
                "- ig.feed.user_feed(user_id) → list",
                "- ig.friendships.follow/unfollow(user_id)",
                "- ig.friendships.get_followers/get_following(user_id)",
                "- ig.direct.send_text(thread_id, text)",
                "- ig.media.get_info(media_pk)",
                "",
                "User model: .username, .followers, .following, .posts_count, .biography",
            ])

        return "\n".join(parts)

    def _build_mode_info(self) -> str:
        """Build detailed mode-specific system prompt addition."""
        parts = ["\n\n# CURRENT SESSION STATUS"]

        if self._mode == "anonymous" or not self._is_logged_in:
            parts.append("""
## MODE: ANONYMOUS (No Login)
- `ig` is available, but no cookie/session
- Only `ig.public.*` methods work
- `ig.users.*`, `ig.feed.*`, `ig.friendships.*` — WILL NOT WORK!

### Rules:
1. ALWAYS use `ig.public.get_profile(username)` — it returns a dict
2. When a dict is returned, use `.get()`: `profile.get('follower_count', 0)`
3. NEVER use `ig.users.get_by_username()` — it requires a session!
4. Follow/unfollow/DM/upload — NOT POSSIBLE, tell the user "login required"
5. If the user requests a task that requires login:
   "This task requires login. Add SESSION_ID, CSRF_TOKEN, DS_USER_ID to the .env file."

### Example pattern (ANONYMOUS):
```python
try:
    profile = ig.public.get_profile('username')
    if profile:
        print(f"Username: {profile.get('username', 'N/A')}")
        print(f"Followers: {profile.get('edge_followed_by', {}).get('count', 0):,}")
        print(f"Following: {profile.get('edge_follow', {}).get('count', 0):,}")
        print(f"Posts: {profile.get('edge_owner_to_timeline_media', {}).get('count', 0):,}")
        print(f"Bio: {profile.get('biography', '')}")
        print(f"Verified: {profile.get('is_verified', False)}")
        # Cache the data
        _cache[profile.get('username', '')] = profile
    else:
        print("User not found")
except Exception as e:
    print(f"Error: {e}")
```

### NOT AVAILABLE IN ANONYMOUS MODE:
- ❌ ig.users.get_by_username() — requires session
- ❌ ig.friendships.follow/unfollow — requires login
- ❌ ig.direct.send_text — requires login
- ❌ ig.media.like/comment — requires login
- ❌ ig.upload.* — requires login
- ❌ ig.account.* — requires login
- ❌ ig.stories.get_tray — requires login

### AVAILABLE IN ANONYMOUS MODE:
- ✅ ig.public.get_profile(username) → dict
- ✅ ig.public.get_posts(username) → list
- ✅ ig.public.get_post_by_url(url) → dict
""")
        elif self._mode == "sync" and self._is_logged_in:
            parts.append("""
## MODE: LOGGED IN (Session available ✅)
- All API methods are available
- ig.users.*, ig.feed.*, ig.friendships.*, ig.direct.* — everything!

### Strategy:
1. First use `ig.users.get_by_username()` — returns User model
2. If unsuccessful, use `ig.public.get_profile()` as fallback
3. Save retrieved data to `_cache` — retrieve from cache on next request

### Example pattern (LOGIN):
```python
try:
    # Check cache first
    if 'username' in _cache:
        user = _cache['username']
        print(f"(From cache) Followers: {user.followers:,}")
    else:
        user = ig.users.get_by_username('username')
        print(f"Username: {user.username}")
        print(f"Followers: {user.followers:,}")
        print(f"Following: {user.following:,}")
        print(f"Posts: {user.posts_count}")
        _cache[user.username] = user  # Save to cache
except Exception as e:
    # Fallback to public API
    try:
        profile = ig.public.get_profile('username')
        count = profile.get('edge_followed_by', {}).get('count', 0)
        print(f"Followers: {count:,}")
        _cache['username'] = profile
    except Exception as e2:
        print(f"Error: {e2}")
```

### ADDITIONAL CAPABILITIES IN LOGIN MODE:
- ✅ ig.friendships.follow/unfollow
- ✅ ig.direct.send_text
- ✅ ig.media.like/comment
- ✅ ig.upload.post_photo/video/reel
- ✅ ig.account.get_current_user
- ✅ ig.stories.get_tray
""")
        elif self._mode == "async":
            parts.append("""
## MODE: ASYNC (AsyncInstagram)
- All methods require `await`
- DO NOT use `import asyncio` and `asyncio.run()`!

### Example:
```python
user = await ig.users.get_by_username('username')
print(f"Followers: {user.followers:,}")
```
""")

        # Add cache info
        parts.append("""
# DATA CACHING
- `_cache` dict — store retrieved data here
- `_is_logged_in` — True/False, check login status
- Always check the cache first
Example: if 'cristiano' in _cache: user = _cache['cristiano']
""")

        return "\n".join(parts)

    def _build_api_reference(self) -> str:
        """Dynamically introspects the ig object to build a live API reference.

        This ensures the agent always knows exactly what modules and methods
        are available at runtime, even if the library is updated.
        """
        if self._ig is None:
            return ""

        import inspect

        parts = ["\n\n# LIVE API REFERENCE (auto-discovered from your ig instance)"]
        parts.append("The following modules and methods are available on the `ig` object right now:\n")

        # Known sub-module attribute names to scan
        _skip = {
            "_session_mgr", "_client", "_anon", "_log",
            "__class__", "__dict__", "__weakref__",
        }

        ig = self._ig
        modules_found = []

        for attr_name in sorted(dir(ig)):
            if attr_name.startswith("_") or attr_name in _skip:
                continue
            try:
                module = getattr(ig, attr_name, None)
            except Exception:
                continue
            if module is None or isinstance(module, (str, int, float, bool, list, dict)):
                continue
            # Must be a sub-module object (has methods)
            if not hasattr(module, "__class__"):
                continue

            methods = []
            for method_name in sorted(dir(module)):
                if method_name.startswith("_"):
                    continue
                try:
                    method = getattr(module, method_name, None)
                except Exception:
                    continue
                if not callable(method):
                    continue

                # Get signature
                try:
                    sig = inspect.signature(method)
                    params = []
                    for pname, param in sig.parameters.items():
                        if pname == "self":
                            continue
                        if param.default is inspect.Parameter.empty:
                            params.append(pname)
                        else:
                            params.append(f"{pname}={param.default!r}")
                    sig_str = f"({', '.join(params)})"
                except (ValueError, TypeError):
                    sig_str = "(...)"

                methods.append(f"  - ig.{attr_name}.{method_name}{sig_str}")

            if methods:
                modules_found.append(f"\n## ig.{attr_name}")
                modules_found.extend(methods)

        if modules_found:
            parts.extend(modules_found)
        else:
            parts.append("(No modules discovered — ig may not be fully initialized)")

        return "\n".join(parts)

    @property
    def is_logged_in(self) -> bool:
        """Check if agent has active login session."""
        return self._is_logged_in

    @property
    def user_cache(self) -> Dict:
        """Access agent's user data cache."""
        return self._user_cache

    # ═══════════════════════════════════════════════════════════
    # PUBLIC API
    # ═══════════════════════════════════════════════════════════

    def register_tool(
        self,
        name: str,
        handler: Callable,
        description: str = "",
        schema: Optional[Dict] = None,
    ):
        """Register a custom plugin tool."""
        self._plugins.register(name, handler, description, schema)

    @property
    def memory(self) -> Optional[AgentMemory]:
        """Access agent memory."""
        return self._memory

    @property
    def cost(self) -> Optional[CostTracker]:
        """Access cost tracker."""
        return self._cost_tracker

    @property
    def plugins(self) -> PluginManager:
        """Access plugin manager."""
        return self._plugins

    @property
    def provider_name(self) -> str:
        """Current provider name."""
        return self._provider.provider_name

    @property
    def mode(self) -> str:
        """Current mode (sync/anonymous/async)."""
        return self._mode

    def ask(
        self,
        message: str,
        step_callback: Optional[Callable[[Dict[str, Any]], None]] = None,
    ) -> AgentResult:
        """
        Ask the agent a question or give a command.

        Args:
            message: Natural language message
            step_callback: Optional callback for real-time step events.
                Receives dicts like {"type": "status", "message": "..."}
                Event types: status, thinking, code, tool_call, tool_result, error

        Returns:
            AgentResult with answer and execution details
        """
        start = time.time()

        if self._verbose:
            _e = emoji('\U0001f916', '[BOT]')
            safe_print(f"\n{_e} Agent working...")

        # Notify: starting
        if step_callback:
            step_callback({"type": "status", "message": "Processing your request..."})

        # Compress history if too long (token optimization)
        self._compress_history()

        # Add user message
        self._history.append({"role": "user", "content": message})

        # Run agent loop
        result = self._agent_loop(step_callback=step_callback)
        result.duration = time.time() - start
        result.tokens_used = self._provider.total_tokens

        if self._verbose and result.success:
            _e = emoji('\u2705', '[OK]')
            safe_print(f"\n{_e} Done ({result.duration:.1f}s, {result.steps} steps)")

        return result

    def chat(self) -> None:
        """
        Interactive chat mode.
        Type 'exit' or 'quit' to exit.
        """
        print("\n" + "\u2550" * 50)
        _e = emoji('\U0001f916', '[BOT]')
        safe_print(f"{_e} InstaAPI Agent \u2014 Interactive mode")
        print(f"   Provider: {self._provider.provider_name}")
        print(f"   Permission: {self._permissions.level.value}")
        print("   'exit' \u2014 quit | 'reset' \u2014 new conversation")
        print("\u2550" * 50)

        while True:
            try:
                user_input = input("\n👤 You: ").strip()
            except (KeyboardInterrupt, EOFError):
                print("\n👋 Goodbye!")
                break

            if not user_input:
                continue
            if user_input.lower() in ("exit", "quit", "q"):
                print("👋 Goodbye!")
                break
            if user_input.lower() in ("reset", "new", "clear"):
                self.reset()
                safe_print(f"{emoji('🔄', '[*]')} Conversation reset")
                continue
            if user_input.lower() in ("history",):
                self._print_history()
                continue

            result = self.ask(user_input)

            if result.success:
                safe_print(f"\n{emoji('🤖', '[BOT]')} Agent: {result.answer}")
            else:
                safe_print(f"\n{emoji('❌', '[ERR]')} Error: {result.error}")

    def reset(self) -> None:
        """Reset conversation history."""
        if self._is_local_provider:
            system_content = self._build_compact_prompt()
        else:
            mode_info = self._build_mode_info()
            api_ref = self._build_api_reference()
            system_content = SYSTEM_PROMPT + mode_info + api_ref
        self._history = [{"role": "system", "content": system_content}]
        self._files_created = []
        self._permissions.reset()
        logger.info("🔄 Agent reset")

    @property
    def history(self) -> List[Dict]:
        """Conversation history."""
        return [m for m in self._history if m["role"] != "system"]

    @property
    def provider_name(self) -> str:
        return self._provider.provider_name

    # ═══════════════════════════════════════════════════════════
    # HISTORY COMPRESSION — TOKEN OPTIMIZATION
    # ═══════════════════════════════════════════════════════════

    # Compress when history has more than this many non-system messages
    _COMPRESS_THRESHOLD = 10
    # Keep this many recent messages after compression
    _KEEP_RECENT = 4

    def _compress_history(self) -> None:
        """
        Compress conversation history to save tokens.

        When history exceeds _COMPRESS_THRESHOLD non-system messages,
        summarize old messages into a single compact entry.
        Keeps: system prompt + summary + last _KEEP_RECENT messages.

        This reduces token usage by ~40-60% on long conversations.
        """
        # Count non-system messages
        non_system = [m for m in self._history if m["role"] != "system"]
        if len(non_system) < self._COMPRESS_THRESHOLD:
            return

        # Split: system + old messages + recent messages
        system_msg = self._history[0]  # Always system
        # Find the split point: keep last _KEEP_RECENT messages
        keep_count = self._KEEP_RECENT
        old_messages = non_system[:-keep_count]
        recent_messages = non_system[-keep_count:]

        if not old_messages:
            return

        # Build summary of old messages
        summary = self._summarize_messages(old_messages)

        if summary:
            # Rebuild history: system + summary + recent
            self._history = [
                system_msg,
                {
                    "role": "user",
                    "content": (
                        f"[CONVERSATION SUMMARY — previous {len(old_messages)} messages]\n"
                        f"{summary}\n"
                        f"[END SUMMARY — continue from here]"
                    ),
                },
                {
                    "role": "assistant",
                    "content": "Understood. I have read the previous conversation summary. Let's continue.",
                },
            ] + recent_messages

            logger.info(
                f"History compressed: {len(non_system)} msgs → "
                f"summary + {len(recent_messages)} recent "
                f"(saved ~{len(old_messages) * 200} tokens)"
            )

    def _summarize_messages(self, messages: List[Dict]) -> str:
        """
        Use LLM to summarize a list of messages into a compact summary.
        Falls back to simple extraction if LLM call fails.
        """
        # Build a compact representation of messages for summarization
        lines = []
        for msg in messages:
            role = msg.get("role", "?")
            content = str(msg.get("content", ""))

            if role == "user":
                lines.append(f"User: {content[:200]}")
            elif role == "assistant":
                # Only include non-empty assistant messages
                if content.strip():
                    lines.append(f"AI: {content[:300]}")
            elif role == "tool":
                # Tool results — very compact
                name = msg.get("name", "tool")
                lines.append(f"[Tool {name}: {content[:150]}]")

        conversation_text = "\n".join(lines)

        # Try to use LLM for smart summarization
        try:
            summary_prompt = [
                {
                    "role": "user",
                    "content": (
                        "Summarize this conversation in 3-5 bullet points. "
                        "Focus on: what the user asked, what data was retrieved, "
                        "what actions were taken. Be very concise.\n\n"
                        f"{conversation_text[:2000]}"
                    ),
                }
            ]

            response = self._provider.generate(
                messages=summary_prompt,
                temperature=0.1,
            )

            if response.content and response.content.strip():
                return response.content.strip()

        except Exception as e:
            logger.warning(f"LLM summarization failed: {e}")

        # Fallback: simple extraction of user messages
        user_msgs = [
            msg.get("content", "")[:100]
            for msg in messages
            if msg.get("role") == "user" and msg.get("content", "").strip()
        ]
        if user_msgs:
            return "Previous requests:\n" + "\n".join(f"- {m}" for m in user_msgs)

        return ""

    # ═══════════════════════════════════════════════════════════
    # AGENT LOOP — THE CORE
    # ═══════════════════════════════════════════════════════════

    def _agent_loop(
        self,
        step_callback: Optional[Callable[[Dict[str, Any]], None]] = None,
    ) -> AgentResult:
        """
        Main agent loop: LLM → Tool → Result → LLM → ...

        Continues until LLM responds without tool calls or max_steps reached.
        """
        result = AgentResult()

        def _emit(event: Dict[str, Any]):
            """Safely emit a step event via the callback."""
            if step_callback:
                try:
                    step_callback(event)
                except Exception:
                    pass  # Never let callback errors break the loop

        for step in range(self._max_steps):
            result.steps = step + 1

            if self._verbose:
                safe_print(f"  {emoji('📍', '>')} Step {step + 1}...", end="")

            # Emit: thinking
            _emit({"type": "thinking", "step": step + 1, "message": f"Step {step + 1}: Analyzing..."})

            # Call LLM
            try:
                response = self._provider.generate(
                    messages=self._history,
                    temperature=0.1,
                )
            except Exception as e:
                result.error = f"Error communicating with AI: {e}"
                logger.error(result.error)
                _emit({"type": "error", "message": result.error})
                return result

            # No tool calls — check if text contains code that should be executed
            if not response.has_tool_calls:
                content = response.content or ""

                # AUTO-EXECUTE: If response has ```python code blocks,
                # extract and run them automatically (MALFORMED fallback scenario)
                extracted_code = self._extract_code_from_text(content)
                if extracted_code:
                    if self._verbose:
                        print(f" 🔧 auto-exec")

                    # Emit: code extracted
                    _emit({"type": "code", "content": extracted_code, "description": "Auto-extracted Python code"})

                    self._history.append({"role": "assistant", "content": content})

                    # Execute the extracted code
                    _emit({"type": "tool_call", "name": "run_instaapi_code", "arguments": {"description": "auto-extracted"}})
                    exec_result = self._handle_code_execution(
                        {"code": extracted_code, "description": "auto-extracted"},
                        step_callback=step_callback,
                    )

                    # Track results
                    result.code_executed = extracted_code
                    if isinstance(exec_result, ExecutionResult):
                        result.execution_result = exec_result

                    # Add tool-like result to history so LLM sees the output
                    result_str = str(exec_result)
                    if len(result_str) > 3000:
                        result_str = result_str[:3000] + "\n... (truncated)"

                    # Emit: tool result
                    _emit({"type": "tool_result", "name": "run_instaapi_code", "output": result_str[:2000],
                           "success": isinstance(exec_result, ExecutionResult) and exec_result.success})

                    self._history.append({
                        "role": "tool",
                        "name": "run_instaapi_code",
                        "content": result_str,
                    })

                    # Continue loop — LLM will see the code output
                    # and generate a proper final answer
                    continue

                # Normal final answer (no code blocks found)
                result.answer = content
                self._history.append({"role": "assistant", "content": content})

                if self._verbose:
                    print(" ✅")
                return result

            # If response has BOTH text content AND tool calls (Gemini behavior)
            # — save the text as potential answer
            if response.content and response.content.strip():
                result.answer = response.content

            if self._verbose:
                print(f" 🔧 {len(response.tool_calls)} tool")

            self._add_assistant_message(response)

            for tc in response.tool_calls:
                # Emit: tool call
                _emit({"type": "tool_call", "name": tc.name, "arguments": tc.arguments})

                tool_result = self._execute_tool(tc, step_callback=step_callback)

                # Add tool result to history
                self._add_tool_result(tc, tool_result)

                # Emit: tool result
                tool_output = str(tool_result)
                _emit({"type": "tool_result", "name": tc.name, "output": tool_output[:2000],
                       "success": not tool_output.startswith("Error")})

                # Track code and files
                if tc.name == "run_instaapi_code":
                    result.code_executed = tc.arguments.get("code", "")
                    if isinstance(tool_result, ExecutionResult):
                        result.execution_result = tool_result
                elif tc.name in ("save_to_file", "create_chart"):
                    filename = tc.arguments.get("filename", "")
                    if filename:
                        result.files_created.append(filename)
                        self._files_created.append(filename)
                elif tc.name == "download_media":
                    output_dir = tc.arguments.get("output_dir", "downloads")
                    result.files_created.append(output_dir)

        # Max steps reached — use last captured answer if available
        if not result.answer.strip():
            result.answer = "Warning: step limit reached. Result may be incomplete."
        return result

    # ═══════════════════════════════════════════════════════════
    # CODE EXTRACTION
    # ═══════════════════════════════════════════════════════════

    @staticmethod
    def _extract_code_from_text(text: str) -> str:
        """
        Extract Python code from markdown code blocks in LLM text.

        Looks for ```python ... ``` blocks and returns the code.
        Returns empty string if no code block found.
        """
        if not text or "```python" not in text:
            return ""

        import re
        # Find all ```python ... ``` blocks
        pattern = r'```python\s*\n(.*?)```'
        matches = re.findall(pattern, text, re.DOTALL)

        if not matches:
            return ""

        # Use the largest code block (most likely the main one)
        code = max(matches, key=len).strip()

        # Sanity check: code should be meaningful (not just comments)
        code_lines = [l for l in code.split('\n') if l.strip() and not l.strip().startswith('#')]
        if len(code_lines) < 1:
            return ""

        return code

    # ═══════════════════════════════════════════════════════════
    # TOOL EXECUTION
    # ═══════════════════════════════════════════════════════════

    def _execute_tool(
        self,
        tool_call: ToolCall,
        step_callback: Optional[Callable[[Dict[str, Any]], None]] = None,
    ) -> Any:
        """Execute a single tool call with permission check."""
        name = tool_call.name
        args = tool_call.arguments

        # Core tools (handled directly)
        if name == "run_instaapi_code":
            return self._handle_code_execution(args, step_callback=step_callback)
        elif name == "save_to_file":
            return self._handle_save_file(args)
        elif name == "ask_user":
            return self._handle_ask_user(args)

        # Extended tools (handled by tools.py)
        elif name in TOOL_HANDLERS:
            handler = TOOL_HANDLERS[name]

            if self._verbose:
                print(f"    🔧 {name}")

            # Some tools need the ig instance
            if name == "download_media":
                if not self._permissions.check(
                    "download.media",
                    f"Download: {args.get('url', '?')}"
                ):
                    return "❌ Permission denied by user"
                return handler(args, ig=self._ig)

            elif name == "http_request":
                if not self._permissions.check(
                    "http.request",
                    f"HTTP {args.get('method', 'GET')} {args.get('url', '?')}"
                ):
                    return "❌ Permission denied by user"
                return handler(args)

            else:
                return handler(args)

        else:
            return f"Unknown tool: {name}"

    def _handle_code_execution(
        self,
        args: Dict,
        step_callback: Optional[Callable[[Dict[str, Any]], None]] = None,
    ) -> str:
        """Handle run_instaapi_code tool call."""
        code = args.get("code", "")
        description = args.get("description", "Code execution")

        if not code.strip():
            return "Error: empty code"

        # Permission check
        if not self._permissions.check_code_execution(code):
            return "Permission denied by user"

        if self._verbose:
            _e = emoji('\U0001f4bb', '>')
            safe_print(f"    {_e} Executing code: {description}")

        # Emit: code about to be executed
        if step_callback:
            try:
                step_callback({"type": "code", "content": code, "description": description})
            except Exception:
                pass

        # Execute in sandbox
        exec_result = self._executor.run(code)

        if exec_result.success:
            output = str(exec_result)
            logger.info(f"Code executed successfully ({exec_result.duration:.2f}s)")
            return output
        else:
            logger.warning(f"Code error: {exec_result.error}")
            return f"Error: {exec_result.error}"

    def _handle_save_file(self, args: Dict) -> str:
        """Handle save_to_file tool call."""
        filename = args.get("filename", "")
        content = args.get("content", "")

        if not filename:
            return "Error: filename not specified"

        # Permission check
        if not self._permissions.check(
            "export.save_file",
            f"Save to file: {filename} ({len(content)} chars)"
        ):
            return "Permission denied by user"

        try:
            # Security: only relative paths
            if os.path.isabs(filename) or ".." in filename:
                return "Error: only relative paths allowed (current directory only)"

            # Get full absolute path for user feedback
            full_path = os.path.abspath(filename)

            # Ensure parent directory exists
            parent_dir = os.path.dirname(full_path)
            if parent_dir and not os.path.exists(parent_dir):
                os.makedirs(parent_dir, exist_ok=True)

            ext = os.path.splitext(filename)[1].lower()

            # Excel format — use openpyxl
            if ext in (".xlsx", ".xls"):
                return self._save_as_excel(filename, content, full_path)

            # Text-based formats
            with open(filename, "w", encoding="utf-8") as f:
                f.write(content)

            if self._verbose:
                _e = emoji('\U0001f4e5', '>')
                safe_print(f"    {_e} Saved: {full_path}")

            return f"File saved successfully!\nPath: {full_path}\nSize: {len(content)} chars"
        except Exception as e:
            return f"Error: could not write to file: {e}"

    def _save_as_excel(self, filename: str, content: str, full_path: str) -> str:
        """Save content as Excel file using openpyxl."""
        try:
            import openpyxl
        except ImportError:
            return (
                "Error: openpyxl is not installed. "
                "Install it: pip install openpyxl"
            )

        try:
            import json

            # Try to parse content as JSON (list of dicts or dict)
            data = json.loads(content)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Data"

            if isinstance(data, list) and data:
                if isinstance(data[0], dict):
                    # List of dicts — headers from keys
                    headers = list(data[0].keys())
                    ws.append(headers)
                    for row in data:
                        ws.append([row.get(h, "") for h in headers])
                else:
                    # List of values
                    for item in data:
                        ws.append([item] if not isinstance(item, list) else item)
            elif isinstance(data, dict):
                # Single dict — key/value columns
                ws.append(["Key", "Value"])
                for k, v in data.items():
                    ws.append([str(k), str(v)])

            wb.save(filename)

            if self._verbose:
                _e = emoji('\U0001f4e5', '>')
                safe_print(f"    {_e} Excel saved: {full_path}")

            return f"Excel file saved successfully!\nPath: {full_path}\nRows: {ws.max_row}"

        except json.JSONDecodeError:
            # Content is not JSON — save as plain text in cells
            wb = openpyxl.Workbook()
            ws = wb.active
            for line in content.split("\n"):
                ws.append([line])
            wb.save(filename)
            return f"Excel file saved (text mode)!\nPath: {full_path}\nRows: {ws.max_row}"
        except Exception as e:
            return f"Error saving Excel: {e}"

    def _handle_ask_user(self, args: Dict) -> str:
        """Handle ask_user tool call."""
        question = args.get("question", "")
        try:
            answer = input(f"\nAgent asks: {question}\nYour answer: ").strip()
            return answer or "(no response)"
        except (KeyboardInterrupt, EOFError):
            return "(user did not respond)"

    # ═══════════════════════════════════════════════════════════
    # MESSAGE HISTORY MANAGEMENT
    # ═══════════════════════════════════════════════════════════

    def _add_assistant_message(self, response: ProviderResponse) -> None:
        """Add assistant message with tool calls to history."""
        provider_name = self._provider.__class__.__name__

        if "OpenAI" in provider_name or "Compatible" in provider_name:
            from .providers.openai_provider import OpenAIProvider
            msg = OpenAIProvider.format_assistant_with_tools(
                response.content, response.tool_calls
            )
            self._history.append(msg)
        elif "Claude" in provider_name:
            # Claude tool results need tool_use_id
            msg = {"role": "assistant", "content": response.content or ""}
            self._history.append(msg)
        else:
            # Generic format for Gemini and others
            msg = {"role": "assistant", "content": response.content or ""}
            self._history.append(msg)

    def _add_tool_result(self, tool_call: ToolCall, result: Any) -> None:
        """Add tool result to history."""
        result_str = str(result)

        # Truncate very long results
        if len(result_str) > 3000:
            result_str = result_str[:3000] + "\n... (qisqartirildi)"

        provider_name = self._provider.__class__.__name__

        if "OpenAI" in provider_name or "Compatible" in provider_name:
            from .providers.openai_provider import OpenAIProvider
            self._history.append(
                OpenAIProvider.format_tool_result(tool_call.id, result_str)
            )
        elif "Claude" in provider_name:
            # Claude expects tool_result in user message
            self._history.append({
                "role": "tool",
                "tool_use_id": tool_call.id,
                "content": result_str,
            })
        else:
            # For Gemini: send as tool response
            self._history.append({
                "role": "tool",
                "name": tool_call.name,
                "content": result_str,
            })

    def _print_history(self) -> None:
        """Print conversation history."""
        for msg in self._history:
            if msg["role"] == "system":
                continue
            role = {"user": "👤", "assistant": "🤖", "tool": "🔧"}.get(msg["role"], "?")
            content = str(msg.get("content", ""))[:100]
            print(f"  {role} {content}")

    # ═══════════════════════════════════════════════════════════
    # UTILITIES
    # ═══════════════════════════════════════════════════════════

    @staticmethod
    def _resolve_api_key(provider: str) -> Optional[str]:
        """Try to resolve API key from environment."""
        return resolve_api_key(provider)

    def __repr__(self) -> str:
        return (
            f"<InstaAgent provider={self._provider.provider_name} "
            f"permission={self._permissions.level.value} "
            f"history={len(self._history)}>"
        )
